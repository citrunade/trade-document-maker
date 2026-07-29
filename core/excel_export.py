"""
Excel (.xlsx) 单据导出模块
使用 openpyxl 生成结构化表格版单据，字段内容与 PDF 导出（core/pdf_export.py）保持一致，
仅版式为电子表格而非排版页面。
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core import calc

DOC_TITLES = {
    "PI": ("PROFORMA INVOICE", "形式发票"),
    "CI": ("COMMERCIAL INVOICE", "商业发票"),
    "PL": ("PACKING LIST", "装箱单"),
}

HEADER_FILL = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
BOX_FILL = PatternFill(start_color="DCE9F7", end_color="DCE9F7", fill_type="solid")
THIN = Side(style="thin", color="BFBFBF")
BOX_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOLD = Font(bold=True)
WHITE_BOLD = Font(bold=True, color="FFFFFF")


def _set(ws, row, col, value, font=None, fill=None, wrap=False, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if wrap or align:
        cell.alignment = Alignment(wrap_text=wrap, vertical="top", horizontal=align or "left")
    cell.border = BOX_BORDER
    return cell


def export_document(doc: dict, filepath: str) -> str:
    doc_type = doc.get("doc_type", "PI")
    financial = doc_type != "PL"
    currency = doc.get("currency", "USD")
    customer = doc.get("customer_snapshot", {})
    own = doc.get("own_snapshot", {})
    delivery = doc.get("delivery_snapshot", {})
    conditions = doc.get("conditions_snapshot", {})
    banking = doc.get("banking_snapshot", {})
    title_en, title_cn = DOC_TITLES.get(doc_type, DOC_TITLES["PI"])

    wb = Workbook()
    ws = wb.active
    ws.title = doc_type
    for col, w in enumerate([14, 26, 12, 10, 12, 12, 8, 10, 10, 14, 16], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    _set(ws, row, 1, f"{title_en} / {title_cn}", font=Font(bold=True, size=14), align="center")
    row += 2

    # 左：收件方（客户）plain 信息
    _set(ws, row, 1, customer.get("name_en", ""), font=BOLD)
    row += 1
    if customer.get("name_cn"):
        _set(ws, row, 1, customer.get("name_cn", "")); row += 1
    if customer.get("address_en"):
        _set(ws, row, 1, customer.get("address_en", "")); row += 1
    if customer.get("tel_phone"):
        _set(ws, row, 1, f"Phone: {customer.get('tel_phone')}"); row += 1
    if customer.get("company_reg_no"):
        _set(ws, row, 1, f"Company Reg. No.: {customer.get('company_reg_no')}"); row += 1
    if customer.get("gst_no"):
        _set(ws, row, 1, f"GST No.: {customer.get('gst_no')}"); row += 1

    # 右：Own plain 信息（与左侧收件方信息并排显示，从固定行开始）
    own_row = 3
    _set(ws, own_row, 7, own.get("company_name_en", ""), font=BOLD); own_row += 1
    if own.get("company_name_cn"):
        _set(ws, own_row, 7, own.get("company_name_cn", "")); own_row += 1
    for line in (own.get("address", "") or "").splitlines():
        if line.strip():
            _set(ws, own_row, 7, line.strip()); own_row += 1
    if own.get("phone"):
        _set(ws, own_row, 7, f"Phone: {own.get('phone')}"); own_row += 1
    if own.get("company_reg_no"):
        _set(ws, own_row, 7, f"Company Reg. No.: {own.get('company_reg_no')}"); own_row += 1
    if own.get("gst_no"):
        _set(ws, own_row, 7, f"GST No.: {own.get('gst_no')}"); own_row += 1

    row = max(row, own_row) + 1

    # Delivery Address (左) + Conditions (左，下方) / Information (右)
    incoterm_line = conditions.get("incoterm", "")
    if incoterm_line and doc.get("destination"):
        incoterm_line = f"{incoterm_line}  {doc.get('destination')}"

    r = row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    _set(ws, r, 1, "Delivery Address", font=BOLD, fill=BOX_FILL)
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    _set(ws, r, 1, delivery.get("company_name", ""), wrap=True)
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    _set(ws, r, 1, delivery.get("address", ""), wrap=True)
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    _set(ws, r, 1, "Conditions", font=BOLD, fill=BOX_FILL)
    r += 1
    for label, value in (
        ("Terms of Payment", conditions.get("terms_of_payment", "")),
        ("Incoterms", incoterm_line),
        ("Shipment by", conditions.get("shipment_by", "")),
    ):
        _set(ws, r, 1, label, font=BOLD)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        _set(ws, r, 2, value, wrap=True)
        r += 1
    left_end = r

    info_rows = [
        ("Doc No", doc.get("doc_number", "")),
        ("Document Date", doc.get("date", "")),
        ("Validity Start Date", doc.get("validity_start", "")),
        ("Validity End Date", doc.get("validity_end", "")),
        ("", ""),
        ("Our Contact", own.get("contact_person", "")),
        ("Telephone", own.get("telephone", "")),
        ("Email", own.get("email", "")),
    ]
    r2 = row
    ws.merge_cells(start_row=r2, start_column=7, end_row=r2, end_column=11)
    _set(ws, r2, 7, "Information", font=BOLD, fill=BOX_FILL)
    r2 += 1
    for label, value in info_rows:
        if not label:
            r2 += 1
            continue
        _set(ws, r2, 7, label, font=BOLD)
        ws.merge_cells(start_row=r2, start_column=8, end_row=r2, end_column=11)
        _set(ws, r2, 8, value, wrap=True)
        r2 += 1
    right_end = r2

    row = max(left_end, right_end) + 2

    # 明细表
    totals = calc.compute_totals(doc.get("lines", []))
    if financial:
        headers = ["No.", "Description EN", "Description CN", "Qty", "Unit", "Unit Price",
                   "Total Price", "COO", "Net Weight", "Total N.W.", "HS Code", "Remark"]
    else:
        headers = ["No.", "Description EN", "Description CN", "Qty", "Unit",
                   "COO", "Net Weight", "Total N.W.", "HS Code", "Remark"]
    for c, h in enumerate(headers, start=1):
        _set(ws, row, c, h, font=WHITE_BOLD, fill=HEADER_FILL, align="center")
    header_row = row
    row += 1

    for i, line in enumerate(totals["lines"], start=1):
        c = 1
        _set(ws, row, c, i); c += 1
        _set(ws, row, c, line.get("name_en", "")); c += 1
        _set(ws, row, c, line.get("name_cn", "")); c += 1
        _set(ws, row, c, line.get("quantity", 0)); c += 1
        _set(ws, row, c, line.get("unit", "")); c += 1
        if financial:
            _set(ws, row, c, round(line.get("unit_price", 0), 2)); c += 1
            _set(ws, row, c, round(line.get("subtotal", 0), 2)); c += 1
        _set(ws, row, c, line.get("coo", "")); c += 1
        _set(ws, row, c, round(line.get("net_weight", 0), 2)); c += 1
        _set(ws, row, c, round(line.get("total_net_weight", 0), 2)); c += 1
        _set(ws, row, c, line.get("hs_code", "")); c += 1
        _set(ws, row, c, line.get("remark", "")); c += 1
        row += 1

    row += 1
    if financial:
        _set(ws, row, 1, f"Total Amount: {currency} {totals['total_amount']:,.2f}", font=BOLD)
        row += 1
        _set(ws, row, 1, calc.amount_in_words(totals["total_amount"], currency))
        row += 1

    _set(ws, row, 1, f"Total Qty: {totals['total_quantity']:g}    Total N.W.: {totals['total_net_weight']:.2f} kg")
    row += 2

    if doc.get("remark"):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        _set(ws, row, 1, "Note", font=BOLD, fill=BOX_FILL)
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        _set(ws, row, 1, doc.get("remark", ""), wrap=True)
        row += 2

    beneficiary_rows = [
        ("Beneficiary Name", banking.get("beneficiary_name", "")),
        ("Beneficiary Bank Name", banking.get("beneficiary_bank_name", "")),
        ("Swift Code", banking.get("beneficiary_swift_code", "")),
        ("Beneficiary Bank's Correspondent Bank", banking.get("correspondent_bank", "")),
        ("Beneficiary Bank Address", banking.get("beneficiary_bank_address", "")),
        ("Beneficiary Account No.", banking.get("beneficiary_account_no", "")),
    ]
    if any(v for _, v in beneficiary_rows):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        _set(ws, row, 1, "Beneficiary Bank Details / 收款银行详情", font=BOLD, fill=BOX_FILL)
        row += 1
        for label, value in beneficiary_rows:
            if not value:
                continue
            _set(ws, row, 1, label, font=BOLD)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=len(headers))
            _set(ws, row, 2, value, wrap=True)
            row += 1

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
    bank_line = f"Bank: {banking.get('bank_name', '')}   Account No.: {banking.get('account_no', '')}   Swift Code: {banking.get('swift_code', '')}"
    _set(ws, row, 1, bank_line, align="center")

    wb.save(filepath)
    return filepath
